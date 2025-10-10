import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, Canvas, Toplevel, Label, Frame
import pandas as pd
import re
import os
import subprocess
import sys
import threading
from datetime import datetime # Importado para a data no nome do arquivo

# Tentar importar openpyxl e seus componentes necessários
try:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill # PatternFill adicionado para o destaque
except ImportError:
    messagebox.showerror("Dependência Faltando",
                         "A biblioteca 'openpyxl' é necessária para formatação avançada do Excel. "
                         "Por favor, instale-a com 'pip install openpyxl' e tente novamente.")
    sys.exit(1)

# Tentar importar pdfplumber
try:
    import pdfplumber
except ImportError:
    messagebox.showerror("Dependência Faltando",
                         "A biblioteca 'pdfplumber' é necessária para extrair dados de PDFs. "
                         "Por favor, instale-a com 'pip install pdfplumber' e tente novamente.")
    sys.exit(1)

# Tentar importar PyMuPDF (fitz) para a nova funcionalidade
try:
    import fitz # PyMuPDF
except ImportError:
    messagebox.showerror("Dependência Faltando",
                         "A biblioteca 'PyMuPDF' (fitz) é necessária para a nova funcionalidade de verificação de 'Valor Cobrado'. "
                         "Por favor, instale-a com 'pip install PyMuPDF' e tente novamente.")
    sys.exit(1)


# --- Funções de Extração e Processamento (Existente) ---

def parse_value(value_str):
    """Converte uma string de valor monetário para float."""
    if not value_str or not isinstance(value_str, str):
        return 0.0
    cleaned_str = value_str.replace('.', '').replace(',', '.')
    try:
        return float(cleaned_str)
    except ValueError:
        return 0.0

def extract_uc_from_block(text_block):
    """Extrai o número da Unidade Consumidora (UC) do bloco de texto."""
    match = re.search(r"(?:UC:|Unidade Consumidora:)\s*(\d+)", text_block)
    if match:
        return match.group(1)
    return None

def extract_valor_total_fatura_from_block(text_block):
    """
    Extrai o valor total da fatura (que será o Valor Líquido) do bloco de texto.
    Tenta encontrar "Valor: R$ XXX" ou "Valor: XXX".
    """
    if not text_block or not isinstance(text_block, str):
        return 0.0

    # Tenta encontrar o padrão "Valor: R$ [valor]"
    # O flag re.DOTALL permite que o '.' corresponda a quebras de linha, caso o valor esteja em outra linha.
    # O flag re.IGNORECASE ignora maiúsculas/minúsculas.
    match_valor_com_simbolo = re.search(r"Valor:\s*R\$\s*([\d\.,]+)", text_block, re.DOTALL | re.IGNORECASE)
    if match_valor_com_simbolo:
        return parse_value(match_valor_com_simbolo.group(1))

    # Se o padrão com R$ não for encontrado, tenta encontrar o padrão "Valor: [valor]" (sem R$)
    # Este padrão é mais genérico e captura números logo após "Valor:", assumindo que não há "R$" se o anterior falhou.
    match_valor_sem_simbolo = re.search(r"Valor:\s*([\d\.,]+)", text_block, re.DOTALL | re.IGNORECASE)
    if match_valor_sem_simbolo:
        return parse_value(match_valor_sem_simbolo.group(1))

    # Se nenhum dos padrões "Valor:" for encontrado, tenta usar o fallback "TOTAL A PAGAR"
    match_total_a_pagar = re.search(r"TOTAL A PAGAR\s*R\$\s*([\d\.,]+)", text_block, re.IGNORECASE)
    if match_total_a_pagar:
        return parse_value(match_total_a_pagar.group(1))

    # Se nada for encontrado, retorna 0.0
    return 0.0

def extract_item_value_from_block(text_block, item_name_pattern):
    """
    Extrai o valor da coluna 'Valor (R$)' para um item específico da seção 'Itens da Fatura'.
    Modificado para pegar o valor na 3ª coluna numérica após o nome do item,
    para lidar com o layout específico dos 'Tributos Retidos'.
    """
    if not text_block or not isinstance(text_block, str):
        return 0.0

    cleaned_text_block = "\n".join(line.strip() for line in text_block.splitlines() if line.strip())
    cleaned_text_block = re.sub(r'[ \t]+', ' ', cleaned_text_block)

    pattern = rf"{item_name_pattern}.*?\s+[\d\.,]+.*?\s+[\d\.,]+.*?\s+(-?[\d\.,]+)"
    match = re.search(pattern, cleaned_text_block, re.MULTILINE | re.IGNORECASE | re.DOTALL)

    if match:
        return parse_value(match.group(1))
    return 0.0

def extract_new_controle_data(text_block):
    """
    Extrai os dados de Energia e Retenção baseados na alíquota de IRPJ (1,2% ou 4,8%)
    para a nova planilha de 'Controle'.
    """
    data = {
        "Energia (1,2%)": 0.0,
        "Retenção(1,2%)": 0.0,
        "Energia (4,8%)": 0.0,
        "Retenção(4,8%)": 0.0
    }

    # Procura a seção de "Itens da Fatura" para focar a extração
    match_itens = re.search(r"Itens da Fatura.*?(?=Valores Medidos|Tributo Retido IRPJ|$)", text_block, re.DOTALL | re.IGNORECASE)
    if not match_itens:
        return data

    relevant_text = match_itens.group(0)
    lines = relevant_text.split('\n')
    
    for line in lines:
        # Verifica se a linha parece conter a estrutura de dados esperada
        if re.search(r'\s1,2\s', line) or re.search(r'\s4,8\s', line):
            # Extrai todos os números (incluindo negativos e decimais com vírgula/ponto) da linha
            numbers = re.findall(r'-?[\d\.,]+', line)
            
            if not numbers:
                continue

            try:
                # Localiza a alíquota (1,2 ou 4,8) na lista de números extraídos
                percent_index = -1
                if '1,2' in numbers:
                    percent_index = numbers.index('1,2')
                    target_key_energia = "Energia (1,2%)"
                    target_key_retencao = "Retenção(1,2%)"
                elif '4,8' in numbers:
                    percent_index = numbers.index('4,8')
                    target_key_energia = "Energia (4,8%)"
                    target_key_retencao = "Retenção(4,8%)"
                else:
                    continue

                # O layout esperado é: [..., Valor (R$), ICMS (R$), Alíquota (%), IRPJ, PIS, COFINS, CSLL]
                # Valor (R$) é o 3º número antes da alíquota
                # Os 4 valores de retenção são os 4 números após a alíquota
                if percent_index >= 3 and (percent_index + 4) < len(numbers):
                    # O 'Valor (R$)' é o 3º valor antes da alíquota, conforme a estrutura de colunas
                    valor_energia = parse_value(numbers[percent_index - 3])
                    
                    # Os valores de retenção (IRPJ, PIS, COFINS, CSLL) são os quatro após a alíquota
                    irpj_val = parse_value(numbers[percent_index + 1])
                    pis_val = parse_value(numbers[percent_index + 2])
                    cofins_val = parse_value(numbers[percent_index + 3])
                    csll_val = parse_value(numbers[percent_index + 4])
                    
                    soma_retencao = abs(irpj_val) + abs(pis_val) + abs(cofins_val) + abs(csll_val)
                    
                    # Soma os valores encontrados aos totais
                    data[target_key_energia] += valor_energia
                    data[target_key_retencao] += soma_retencao
            
            except (ValueError, IndexError):
                # Ocorre se o formato da linha for inesperado. Ignora a linha e continua.
                continue
    
    return data


def extract_fatura_data_from_text_block(text_block, df_base, pdf_filename_for_error_logging, logger_func, page_num=None):
    """
    Extrai todos os dados de uma fatura a partir de um bloco de texto.
    Retorna um dicionário com os dados ou um dicionário de erro.
    """
    uc_number = extract_uc_from_block(text_block)
    if not uc_number:
        return None

    base_info = df_base[df_base['UC'].astype(str) == uc_number]
    if base_info.empty:
        error_msg = f"UC {uc_number} (de {pdf_filename_for_error_logging}) não encontrada na planilha base."
        if logger_func:
            logger_func(error_msg, "ERROR")
        return {"error": error_msg, "UC": uc_number, "Numero da Pagina": pdf_filename_for_error_logging}

    cod_reg = base_info['Cod de Reg'].iloc[0]
    nome_base = base_info['Nome'].iloc[0]

    valor_liquido_fatura = extract_valor_total_fatura_from_block(text_block)
    if valor_liquido_fatura == 0.0 and logger_func:
         logger_func(f"Valor Líquido da fatura (Valor Total da Fatura) não encontrado ou zerado para UC {uc_number} em {pdf_filename_for_error_logging}. Verifique o PDF ou o padrão de extração.", "WARNING")

    tributos_retidos_patterns = {
        "IRPJ": r"Tributo Retido IRPJ",
        "PIS": r"Tributo Retido PIS",
        "COFINS": r"Tributo Retido COFINS",
        "CSLL": r"Tributo Retido CSLL"
    }

    soma_valores_negativos_tributos = 0.0
    found_any_tax_value_non_zero = False

    for nome_tributo, pattern_str in tributos_retidos_patterns.items():
        valor_tributo = extract_item_value_from_block(text_block, re.escape(pattern_str))
        soma_valores_negativos_tributos += valor_tributo
        if valor_tributo != 0.0:
            found_any_tax_value_non_zero = True

    retencao_tributos = abs(soma_valores_negativos_tributos)

    if retencao_tributos == 0.0 and not found_any_tax_value_non_zero and logger_func:
         logger_func(f"INFO: Nenhum item de tributo retido ('Tributo Retido IRPJ/PIS/COFINS/CSLL') encontrado ou extraído com valor não zero para UC {uc_number} em {pdf_filename_for_error_logging}. 'RETENÇÃO (R$)' será 0.00.", "INFO")

    cosip_item_name_pattern = r"COSIP Municipal"
    valor_cosip = extract_item_value_from_block(text_block, cosip_item_name_pattern)

    if valor_cosip == 0.0 and logger_func:
        logger_func(f"INFO: COSIP (ou 'COSIP Municipal') não encontrado ou extraído com valor zero para UC {uc_number} em {pdf_filename_for_error_logging}. 'COSIP (R$)' será 0.00.", "INFO")

    valor_bruto_fatura_calculado = valor_liquido_fatura + retencao_tributos
    valor_energia_calculado = valor_bruto_fatura_calculado - valor_cosip

    numero_pagina_display = f"{pdf_filename_for_error_logging} (Pág. {page_num + 1})" if page_num is not None else pdf_filename_for_error_logging

    # Dados para a aba Relatorio (formato antigo)
    fatura_data = {
        "UC": uc_number,
        "Centro de Custo": cod_reg,
        "Subseção": nome_base,
        "ENERGIA (R$)": valor_energia_calculado,
        "COSIP (R$)": valor_cosip,
        "Valor Bruto (R$)": valor_bruto_fatura_calculado,
        "RETENÇÃO (R$)": retencao_tributos,
        "LÍQUIDO (R$)": valor_liquido_fatura,
        "Numero da Pagina": numero_pagina_display
    }

    # Extrai e adiciona os novos dados EXCLUSIVAMENTE para a aba de Controle
    controle_data = extract_new_controle_data(text_block)
    fatura_data.update(controle_data)

    return fatura_data

def process_pdf_file(pdf_path, df_base, logger_func, progress_callback):
    """
    Processa um único arquivo PDF.
    Retorna uma lista de dicionários (dados da fatura ou erros).
    """
    results_for_this_pdf = []
    pdf_filename = os.path.basename(pdf_path)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                error_msg = f"PDF sem páginas: {pdf_filename}"
                logger_func(error_msg, "ERROR")
                results_for_this_pdf.append({"error": error_msg, "Numero da Pagina": pdf_filename})
                if progress_callback:
                    progress_callback(0)
                return results_for_this_pdf

            total_pages_in_pdf = len(pdf.pages)

            for page_num, page in enumerate(pdf.pages):
                page_text = page.extract_text(x_tolerance=2, y_tolerance=3)
                if not page_text or not page_text.strip():
                    logger_func(f"Página {page_num + 1} de {pdf_filename} não contém texto extraível.", "INFO")
                    if progress_callback:
                        progress_callback(1)
                    continue

                uc_pattern = r"(?:UC:|Unidade Consumidora:)\s*\d+"
                matches = list(re.finditer(uc_pattern, page_text))

                if not matches:
                    if page_num == 0:
                         logger_func(f"Nenhuma UC explícita na página {page_num+1} (provável sumário) de {pdf_filename}. Pulando página.", "INFO")
                         if progress_callback:
                            progress_callback(1)
                         continue
                    else:
                        logger_func(f"Nenhuma UC explícita na página {page_num+1} de {pdf_filename}. Tentando processar a página inteira como um bloco único.", "INFO")
                        fatura_data = extract_fatura_data_from_text_block(page_text, df_base, pdf_filename, logger_func, page_num=page_num)
                        if fatura_data:
                            results_for_this_pdf.append(fatura_data)
                        if progress_callback:
                           progress_callback(1)
                        continue

                for i, match in enumerate(matches):
                    start_block = match.start()
                    end_block = matches[i+1].start() if i + 1 < len(matches) else len(page_text)
                    current_text_block = page_text[start_block:end_block]

                    fatura_data = extract_fatura_data_from_text_block(current_text_block, df_base, pdf_filename, logger_func, page_num=page_num)
                    if fatura_data:
                        results_for_this_pdf.append(fatura_data)

                if progress_callback:
                   progress_callback(1)

            if not results_for_this_pdf:
                 no_data_msg = f"Nenhum dado de fatura (com UC identificável) ou erro relevante encontrado em {pdf_filename} após processar todas as páginas com texto extraível."
                 logger_func(no_data_msg, "WARNING")

    except Exception as e:
        critical_error_msg = f"Erro crítico ao processar {pdf_filename}: {e}"
        logger_func(critical_error_msg, "CRITICAL_ERROR")
        results_for_this_pdf.append({"error": critical_error_msg, "Numero da Pagina": pdf_filename, "UC": "N/A"})
        if progress_callback:
           try:
               with fitz.open(pdf_path) as pdf_err: # Using fitz here for consistency if pdfplumber failed to open
                    progress_callback(len(pdf_err.pages) - (page_num + 1 if 'page_num' in locals() else 0))
           except Exception:
                progress_callback(1)

    return results_for_this_pdf


# --- Classe da Interface Gráfica ---
class AppCelescReporter:
    def __init__(self, root_window):
        self.root = root_window
        self.root.title("Gerador de Relatório Celesc - ver 1.2a")
        self.center_window(700, 650)
        self.root.resizable(False, False)

        if getattr(sys, 'frozen', False):
             basedir = os.path.dirname(sys.executable)
        else:
             basedir = os.path.dirname(__file__)

        self.base_sheet_path = os.path.join(basedir, "base", "database.xlsx")

        self.icon_path = os.path.join(basedir, "base", "icon.ico")
        if os.path.exists(self.icon_path):
            try:
                self.root.iconbitmap(self.icon_path)
            except tk.TclError as e:
                print(f"Erro ao carregar ícone: {e}")
        else:
            print(f"Aviso: Arquivo de ícone não encontrado em {self.icon_path}")

        self.df_base = None
        self.pdf_files = []
        self.total_pages_to_process = 0
        self.processed_pages_count = 0
        self.output_dir = os.path.join(os.path.expanduser("~"), "Desktop")

        self.current_severity = 0
        self.SEVERITY_MAP = {
            "INFO": 0, "DEBUG": 0, "SUCCESSO": 0,
            "AVISO": 1,
            "ERRO": 2, "ERRO_CRITICO!": 2
        }

        self.has_specific_warnings = False # Flag para avisos específicos (para o resumo final)
        self.account_values_mismatched = False # Nova flag para a verificação de valores

        style = ttk.Style(self.root)
        style.theme_use('clam')
        style.configure("Default.Horizontal.TProgressbar", troughcolor='white', background='green')
        style.configure("Success.Horizontal.TProgressbar", troughcolor='white', background='green')
        style.configure("Warning.Horizontal.TProgressbar", troughcolor='white', background='yellow')
        style.configure("Error.Horizontal.TProgressbar", troughcolor='white', background='red')

        self.theme_background_color = style.lookup('TFrame', 'background')

        main_frame = ttk.Frame(self.root, padding="10 10 10 10")
        main_frame.pack(expand=True, fill=tk.BOTH)

        # --- 1. Planilha Base de UCs ---
        base_frame = ttk.LabelFrame(main_frame, text="Planilha Base de UCs", padding="10")
        base_frame.pack(fill=tk.X, pady=5)

        self.base_path_label = ttk.Label(base_frame, text=f"Caminho: {self.base_sheet_path}", wraplength=650, cursor="hand2")
        self.base_path_label.pack(fill=tk.X)
        self.base_path_label.bind("<Button-1>", lambda e: self.open_base_sheet_folder())

        self.base_status_label = ttk.Label(base_frame, text="Status: Não carregada")
        self.base_status_label.pack(fill=tk.X)

        # --- Seção Log em Tempo Real ---
        log_frame = ttk.LabelFrame(main_frame, text="Log de Processamento", padding="10")

        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=10, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.tag_config("INFO", foreground="black")
        self.log_text.tag_config("AVISO", foreground="orange")
        self.log_text.tag_config("ERRO", foreground="red")
        self.log_text.tag_config("ERRO_CRITICO", foreground="red", font=('TkDefaultFont', 9, 'bold'))
        self.log_text.tag_config("SUCCESSO", foreground="green")
        self.log_text.tag_config("DEBUG", foreground="gray")

        self.load_base_sheet() # Carrega a planilha base ao iniciar

        # --- 2. Container para PDF e Parâmetros ---
        pdf_params_container_frame = ttk.Frame(main_frame)
        pdf_params_container_frame.pack(fill=tk.X, pady=5)

        # --- 2.1 Arquivos PDF das Faturas (Esquerda) ---
        pdf_frame = ttk.LabelFrame(pdf_params_container_frame, text="Arquivos PDF das Faturas", padding="10")
        pdf_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.pdf_button = ttk.Button(pdf_frame, text="Selecionar PDFs da Celesc", command=self.select_pdfs)
        self.pdf_button.pack(side=tk.LEFT, padx=(0,10))
        self.pdf_label = ttk.Label(pdf_frame, text="Nenhum PDF selecionado")
        self.pdf_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # --- 2.2 Parâmetros (Direita) ---
        params_frame = ttk.LabelFrame(pdf_params_container_frame, text="Parâmetros", padding="10")
        params_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))

        self.gerar_controle_var = tk.BooleanVar()
        self.gerar_txt_var = tk.BooleanVar()
        self.gerar_relatorio_var = tk.BooleanVar(value=True)

        controle_check = ttk.Checkbutton(params_frame, text="Gerar Controle", variable=self.gerar_controle_var)
        controle_check.pack(side=tk.LEFT, padx=1, pady=2)

        # Separador vertical entre opções
        separator_canvas = Canvas(params_frame, width=1, height=15, bg=self.theme_background_color, highlightthickness=0)
        separator_canvas.create_line(0, 0, 0, 15, fill="gray")
        separator_canvas.pack(side=tk.LEFT, padx=(5, 5))

        txt_check = ttk.Checkbutton(params_frame, text="Gerar TXT", variable=self.gerar_txt_var, command=self.on_toggle_gerar_txt)
        txt_check.pack(side=tk.LEFT, padx=5, pady=2)

        # Separador vertical entre opções
        separator_canvas_1 = Canvas(params_frame, width=1, height=15, bg=self.theme_background_color, highlightthickness=0)
        separator_canvas_1.create_line(0, 0, 0, 15, fill="gray")
        separator_canvas_1.pack(side=tk.LEFT, padx=(5, 5))

        self.relatorio_checkbox = ttk.Checkbutton(params_frame, text="Gerar Relatorio", variable=self.gerar_relatorio_var, state=tk.DISABLED)
        self.relatorio_checkbox.pack(side=tk.LEFT, padx=5, pady=2)

        # --- 3. Pasta de Saída do Relatório ---
        output_frame = ttk.LabelFrame(main_frame, text="Pasta de Saída do Relatório", padding="10")
        output_frame.pack(fill=tk.X, pady=5)
        self.output_dir_button = ttk.Button(output_frame, text="Definir Pasta de Saída", command=self.select_output_dir)
        self.output_dir_button.pack(side=tk.LEFT, padx=(0,10))
        self.output_label = ttk.Label(output_frame, text=f"Padrão: {self.output_dir}")
        self.output_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- 4. Log de Processamento ---
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # --- 5. Action Frame ---
        action_frame = ttk.Frame(main_frame, padding="10")
        action_frame.pack(fill=tk.X, pady=10)

        self.progress_bar = ttk.Progressbar(action_frame, orient="horizontal", length=300, mode="determinate",
                                            style="Default.Horizontal.TProgressbar")
        self.progress_bar.pack(pady=5, fill=tk.X)

        self.process_button = ttk.Button(action_frame, text="Iniciar Processamento de Relatório", command=self.start_processing)
        self.process_button.pack(pady=5)

        self.status_label = ttk.Label(action_frame, text="Aguardando configuração...")
        self.status_label.pack(fill=tk.X, pady=5)

        # Botão de Informação "i"
        show_info_button_canvas = create_rounded_button(root, "i", self.show_info, width=20, height=20, bg_color=self.theme_background_color)
        show_info_button_canvas.place(relx=1.0, rely=1.0, x=-10, y=-10, anchor="se")

    def set_progress_bar_style(self, style_name):
        """Define o estilo visual da barra de progresso."""
        try:
            self.progress_bar.config(style=style_name)
        except tk.TclError as e:
            self.log_message(f"Erro ao aplicar estilo '{style_name}' à barra de progresso: {e}. Usando estilo padrão.", "WARNING")
            self.progress_bar.config(style="Default.Horizontal.TProgressbar")

    def log_message(self, message, level="INFO"):
        """Insere uma mensagem no widget de log, configura tags para colorir e atualiza a severidade."""
        display_message = f"[{level}] {message}\n"
        tag = level.upper()

        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, display_message, tag)
        self.log_text.config(state=tk.DISABLED)
        self.log_text.see(tk.END)

        new_severity = self.SEVERITY_MAP.get(level, 0)
        if new_severity > self.current_severity:
            self.current_severity = new_severity
            if self.current_severity == 0:
                self.set_progress_bar_style("Success.Horizontal.TProgressbar")
            elif self.current_severity == 1:
                self.set_progress_bar_style("Warning.Horizontal.TProgressbar")
            else:
                self.set_progress_bar_style("Error.Horizontal.TProgressbar")

        if level == "WARNING" and message.startswith("Valor Líquido da fatura (Valor Total da Fatura) não encontrado ou zerado para UC"):
            self.has_specific_warnings = True

    def update_progress(self, pages_processed):
        """Atualiza a barra de progresso (valor) e o status label."""
        self.processed_pages_count += pages_processed
        current_progress = self.processed_pages_count
        total_steps = self.total_pages_to_process

        if current_progress > total_steps:
             current_progress = total_steps

        self.root.after(0, lambda: self.progress_bar.config(value=current_progress))
        if pages_processed > 0 or current_progress == total_steps:
            self.root.after(0, lambda: self.status_label.config(text=f"Processando página {current_progress}/{total_steps}..."))

    def center_window(self, width, height):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        self.root.geometry(f'{width}x{height}+{int(x)}+{int(y)}')

    def load_base_sheet(self):
        """Carrega a planilha base de UCs e atualiza o status na interface."""
        self.log_message("Tentando carregar planilha base...", "INFO")
        try:
            if not os.path.exists(self.base_sheet_path):
                msg = f"Status: ERRO - Arquivo base não encontrado em {self.base_sheet_path}"
                self.base_status_label.config(text=msg, foreground="red")
                self.log_message(msg, "ERROR")
                self.df_base = None
                return

            self.df_base = pd.read_excel(self.base_sheet_path, engine='openpyxl', dtype={'UC': str, 'Cod de Reg': str, 'Nome': str})
            required_cols = ['UC', 'Cod de Reg', 'Nome']
            if not all(col in self.df_base.columns for col in required_cols):
                missing_cols = [col for col in required_cols if col not in self.df_base.columns]
                msg = f"Status: ERRO - Colunas faltando na planilha base: {', '.join(missing_cols)}. Necessárias: {', '.join(required_cols)}"
                self.base_status_label.config(text=msg, foreground="red")
                self.log_message(msg, "ERROR")
                self.df_base = None
                return

            self.df_base.dropna(subset=['UC'], inplace=True)
            self.df_base['UC'] = self.df_base['UC'].astype(str).str.strip()

            num_ucs = len(self.df_base)
            if num_ucs == 0:
                msg = "Status: Planilha base carregada, mas sem UCs válidas após limpeza."
                self.base_status_label.config(text=msg, foreground="orange")
                self.log_message(msg, "WARNING")
            else:
                msg = f"Status: Planilha base carregada. {num_ucs} UCs encontradas."
                self.base_status_label.config(text=msg, foreground="green")
                self.log_message(msg, "INFO")
        except Exception as e:
            msg = f"Status: ERRO ao carregar planilha base - {e}"
            self.base_status_label.config(text=msg, foreground="red")
            self.log_message(msg, "CRITICAL_ERROR")
            self.df_base = None

    def open_base_sheet_folder(self):
        """Abre o diretório onde a planilha base está localizada."""
        if not os.path.exists(self.base_sheet_path):
            self.log_message(f"Caminho da planilha base não encontrado para abrir: {self.base_sheet_path}", "ERROR")
            messagebox.showerror("Erro", "Arquivo da planilha base não encontrado.")
            return

        folder_path = os.path.dirname(self.base_sheet_path)
        self.log_message(f"Abrindo pasta da planilha base: {folder_path}", "INFO")
        try:
            if sys.platform == "win32":
                os.startfile(folder_path)
            elif sys.platform == "darwin": # macOS
                subprocess.call(("open", folder_path))
            else: # linux variants
                subprocess.call(("xdg-open", folder_path))
        except Exception as e:
            self.log_message(f"Erro ao tentar abrir o diretório '{folder_path}': {e}", "ERROR")
            messagebox.showerror("Erro ao Abrir Pasta", f"Não foi possível abrir a pasta:\n{folder_path}\nErro: {e}")

    def select_pdfs(self):
        """Permite ao usuário selecionar múltiplos arquivos PDF."""
        files = filedialog.askopenfilenames(
            title="Selecione os arquivos PDF da Celesc",
            filetypes=(("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*"))
        )
        if files:
            self.pdf_files = list(files)
            self.pdf_label.config(text=f"{len(self.pdf_files)} PDF(s) selecionado(s)")
            self.log_message(f"{len(self.pdf_files)} PDF(s) selecionado(s).", "INFO")
        else:
            self.pdf_label.config(text="Nenhum PDF selecionado")
            self.log_message("Nenhum PDF selecionado.", "INFO")
            self.pdf_files = []

    def on_toggle_gerar_txt(self):
        """
        Garante que, se 'Gerar TXT' for marcado, 'Gerar Controle' também seja.
        Esta lógica só é ativada ao marcar a caixa, não ao desmarcar.
        """
        if self.gerar_txt_var.get(): # Se o checkbox 'Gerar TXT' está agora marcado
            if not self.gerar_controle_var.get(): # E o checkbox 'Gerar Controle' não está marcado
                self.gerar_controle_var.set(True) # Marca o 'Gerar Controle'

    def select_output_dir(self):
        """Permite ao usuário selecionar a pasta de saída para o relatório."""
        directory = filedialog.askdirectory(title="Selecione a pasta para salvar o relatório")
        if directory:
            self.output_dir = directory
            self.output_label.config(text=self.output_dir)
            self.log_message(f"Pasta de saída definida para: {self.output_dir}", "INFO")

    def start_processing(self):
        """Inicia o processo de extração e geração do relatório em uma nova thread."""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete('1.0', tk.END)
        self.log_text.config(state=tk.DISABLED)

        self.current_severity = 0
        self.has_specific_warnings = False # Reset flag
        self.account_values_mismatched = False # Reset flag

        self.log_message("Iniciando processo de verificação...", "INFO")

        self.load_base_sheet()

        if self.df_base is None or self.df_base.empty:
            msg = "Planilha base de UCs não carregada, inválida ou vazia. Verifique o arquivo 'base/database.xlsx'."
            self.log_message(msg, "ERROR")
            messagebox.showerror("Erro de Configuração", msg)
            self.set_progress_bar_style("Error.Horizontal.TProgressbar")
            self.progress_bar["value"] = 1
            self.progress_bar["maximum"] = 1
            self.status_label.config(text="Erro de configuração: Planilha base.")
            self.process_button.config(state=tk.NORMAL)
            return

        if not self.pdf_files:
            msg = "Nenhum arquivo PDF foi selecionado para processamento."
            self.log_message(msg, "ERROR")
            messagebox.showerror("Erro de Configuração", msg)
            self.set_progress_bar_style("Error.Horizontal.TProgressbar")
            self.progress_bar["value"] = 1
            self.progress_bar["maximum"] = 1
            self.status_label.config(text="Erro de configuração: PDFs não selecionados.")
            self.process_button.config(state=tk.NORMAL)
            return

        if not self.output_dir or not os.path.isdir(self.output_dir):
            msg = "Pasta de saída inválida ou não definida."
            self.log_message(msg, "ERROR")
            messagebox.showerror("Erro de Configuração", msg)
            self.set_progress_bar_style("Error.Horizontal.TProgressbar")
            self.progress_bar["value"] = 1
            self.progress_bar["maximum"] = 1
            self.status_label.config(text="Erro de configuração: Pasta de saída inválida.")
            self.process_button.config(state=tk.NORMAL)
            return

        self.total_pages_to_process = 0
        self.log_message("Contando total de páginas nos PDFs...", "INFO")
        temp_total_pages = 0
        for pdf_path in self.pdf_files:
            try:
                with fitz.open(pdf_path) as pdf: # Usando fitz para contagem mais confiável
                    temp_total_pages += len(pdf)
            except Exception as e:
                self.log_message(f"AVISO: Não foi possível contar páginas em {os.path.basename(pdf_path)}: {e}. Assumindo 1 página para o progresso.", "WARNING")
                temp_total_pages += 1
        self.total_pages_to_process = max(1, temp_total_pages)
        self.log_message(f"Total de páginas a processar: {self.total_pages_to_process}", "INFO")

        self.processed_pages_count = 0
        self.status_label.config(text=f"Preparando para processar {self.total_pages_to_process} páginas...")
        self.process_button.config(state=tk.DISABLED)
        self.progress_bar["value"] = 0
        self.progress_bar["maximum"] = self.total_pages_to_process
        self.set_progress_bar_style("Success.Horizontal.TProgressbar")

        self.root.update_idletasks()

        processing_thread = threading.Thread(target=self._actual_processing_task)
        processing_thread.start()

    def _actual_processing_task(self):
        """Contém o loop principal de processamento de PDF, executa em uma thread separada."""
        all_extracted_data = []
        error_items = []
        erros_encontrados_no_processamento = False

        self.root.after(0, lambda: self.progress_bar.config(value=0, maximum=self.total_pages_to_process))
        self.root.after(0, lambda: self.status_label.config(text=f"Iniciando processamento de {self.total_pages_to_process} páginas..."))
        self.set_progress_bar_style("Success.Horizontal.TProgressbar")

        self.log_message(f"Iniciando processamento de {len(self.pdf_files)} PDFs ({self.total_pages_to_process} páginas totais estimadas)...", "INFO")

        for pdf_path in self.pdf_files:
            pdf_name = os.path.basename(pdf_path)
            self.log_message(f"Processando PDF: {pdf_name}", "INFO")

            results_from_pdf = process_pdf_file(pdf_path, self.df_base, self.log_message, self.update_progress)

            for item in results_from_pdf:
                if isinstance(item, dict):
                    if "error" in item:
                        erros_encontrados_no_processamento = True
                        error_items.append(item)
                    else:
                        all_extracted_data.append(item)
        
        if error_items:
            erros_encontrados_no_processamento = True

        # --- Nova etapa: Extrair e verificar 'Valor Cobrado' para cada PDF ---
        all_valor_cobrado_results = []
        self.log_message("\n--- Iniciando verificação de 'Valor Cobrado' ---", "INFO")
        for pdf_path in self.pdf_files:
            pdf_name = os.path.basename(pdf_path)
            cobrado_val, cobrado_str, liquido_total_verified, status_msgs = self.extract_and_verify_valor_cobrado(pdf_path)

            for msg in status_msgs:
                if "Aviso" in msg:
                    self.log_message(f"[{pdf_name}] {msg}", "WARNING")
                elif "Erro" in msg:
                    self.log_message(f"[{pdf_name}] {msg}", "ERROR")
                else:
                    self.log_message(f"[{pdf_name}] {msg}", "INFO")

            all_valor_cobrado_results.append({"pdf": pdf_name, "valor_cobrado": cobrado_val, "liquido_total_verified": liquido_total_verified})
        self.log_message("--- Verificação de 'Valor Cobrado' concluída ---", "INFO")

        self.root.after(0, lambda: self.progress_bar.config(value=self.total_pages_to_process))
        self.root.after(0, lambda: self.status_label.config(text=f"Processamento concluído! Gerando relatório..."))

        self.root.after(100, lambda: self._processing_complete(all_extracted_data, error_items, erros_encontrados_no_processamento, all_valor_cobrado_results))


    def _processing_complete(self, all_extracted_data, error_items, erros_encontrados_no_processamento, all_valor_cobrado_results):
        """Finaliza o processamento, cria o relatório Excel e atualiza a GUI."""

        # Colunas para a aba 'Relatorio' (sem as novas colunas)
        final_columns_order_data = [
            "UC", "Centro de Custo", "Subseção",
            "ENERGIA (R$)",
            "COSIP (R$)",
            "Valor Bruto (R$)",
            "RETENÇÃO (R$)",
            "LÍQUIDO (R$)",
            "Numero da Pagina"
        ]
        # Colunas para a aba de Erros
        final_columns_order_errors = [
            "UC", "Centro de Custo", "Subseção",
            "ENERGIA (R$)",
            "COSIP (R$)",
            "Valor Bruto (R$)",
            "RETENÇÃO (R$)",
            "LÍQUIDO (R$)",
            "Numero da Pagina",
            "Observação"
        ]

        # Nomes das colunas de moeda para formatação no Excel (apenas para 'Relatorio')
        currency_cols_names_for_excel_fmt = [
            "ENERGIA (R$)",
            "COSIP (R$)",
            "Valor Bruto (R$)",
            "RETENÇÃO (R$)",
            "LÍQUIDO (R$)"
        ]

        # --- Cria o DataFrame completo com todos os dados extraídos ---
        df_full_data = pd.DataFrame(all_extracted_data)

        # --- Geração do nome do arquivo com data ---
        try:
            today_str = datetime.today().strftime("%d.%m.%Y")
            output_filename = f"{today_str} Repasse-Celesc.xlsx"
            output_file_path = os.path.join(self.output_dir, output_filename)
            self.log_message(f"Nome do arquivo de saída gerado: {output_filename}", "INFO")
        except Exception as e:
            self.log_message(f"Erro ao gerar nome do arquivo de saída: {e}. Usando nome padrão.", "WARNING")
            output_file_path = os.path.join(self.output_dir, "Relatorio_Celesc.xlsx")

        # --- PREPARAR DADOS PARA A ABA 'CONTROLE' (SE SOLICITADO) ---
        df_controle = pd.DataFrame()
        if self.gerar_controle_var.get():
            self.log_message("Preparando dados para a aba 'Controle'...", "INFO")
            if not df_full_data.empty and not df_full_data[df_full_data['UC'].notna()].empty:
                # Novas colunas e dicionário de agregação para a aba 'Controle'
                new_controle_cols = [
                    "Energia (1,2%)", "Retenção(1,2%)",
                    "Energia (4,8%)", "Retenção(4,8%)"
                ]
                
                # Garante que as novas colunas existam no dataframe antes de agrupar
                for col in new_controle_cols:
                    if col not in df_full_data.columns:
                        df_full_data[col] = 0.0
                
                # Dicionário de agregação para agrupar por Centro de Custo e Subseção
                controle_agg_dict = {
                    'UC': lambda x: '\n'.join(sorted(x.astype(str).unique())), # Concatena UCs
                    'COSIP (R$)': 'sum'
                }
                # Adiciona as outras colunas numéricas para soma
                for col in new_controle_cols:
                    controle_agg_dict[col] = 'sum'

                # Agrupa por Centro de Custo e Subseção, somando valores e concatenando UCs
                df_controle = df_full_data.groupby(['Centro de Custo', 'Subseção'], as_index=False).agg(controle_agg_dict)
                
                # Reordena as colunas para o formato final especificado
                final_controle_order = [
                    'UC', 'Centro de Custo', 'Subseção',
                    'COSIP (R$)',
                    'Energia (1,2%)', 'Retenção(1,2%)',
                    'Energia (4,8%)', 'Retenção(4,8%)'
                ]
                
                # Filtra para garantir que apenas colunas existentes sejam usadas
                df_controle = df_controle.reindex(columns=final_controle_order)

                # Adiciona a linha de totais à aba 'Controle'
                if not df_controle.empty:
                    # --- INÍCIO DA MODIFICAÇÃO: Geração dos arquivos TXT ---
                    if self.gerar_txt_var.get():
                        try:
                            self.log_message("Iniciando geração de arquivos TXT...", "INFO")
                            # Cria a pasta de saída para os TXTs com base no nome do Excel
                            txt_folder_name = os.path.splitext(output_filename)[0]
                            txt_output_dir = os.path.join(self.output_dir, txt_folder_name)
                            os.makedirs(txt_output_dir, exist_ok=True)
                            self.log_message(f"Pasta para arquivos TXT criada em: {txt_output_dir}", "INFO")
                    
                            # Mapeamento de nome de arquivo para coluna de dados
                            txt_map = {
                                "Rateio Cosip.txt": "COSIP (R$)",
                                "Rateio Energia 1.2.txt": "Energia (1,2%)",
                                "Rateio Energia 4.8.txt": "Energia (4,8%)"
                            }
                    
                            # Itera sobre o mapa para gerar cada arquivo TXT
                            for filename, data_column in txt_map.items():
                                txt_file_path = os.path.join(txt_output_dir, filename)
                                lines_to_write = []
                    
                                # Itera sobre as linhas do DataFrame 'Controle' (antes de adicionar totais)
                                for index, row in df_controle.iterrows():
                                    centro_custo = row['Centro de Custo']
                                    value = row[data_column]
                    
                                    # Processa apenas se o valor não for nulo e for maior que zero
                                    if pd.notna(value) and abs(value) > 1e-9:
                                        # Formatação do valor numérico
                                        formatted_value = f"{value:.2f}".replace('.', ',')
                                        if formatted_value.endswith(",00"):
                                            formatted_value = formatted_value[:-3]
                                        
                                        if pd.notna(centro_custo) and str(centro_custo).strip():
                                            lines_to_write.append(f"{centro_custo}#SEP#{formatted_value}")
                                
                                # Escreve as linhas no arquivo
                                if lines_to_write:
                                    with open(txt_file_path, 'w', encoding='utf-8') as f:
                                        f.write('\n'.join(lines_to_write))
                                    self.log_message(f"Arquivo '{filename}' gerado com {len(lines_to_write)} linhas.", "SUCCESSO")
                                else:
                                    self.log_message(f"Nenhum dado válido para gerar o arquivo '{filename}'.", "INFO")
                        
                        except Exception as e:
                            self.log_message(f"Erro CRÍTICO ao gerar arquivos TXT: {e}", "ERRO_CRITICO")
                            messagebox.showerror("Erro na Geração de TXT", f"Ocorreu um erro ao gerar os arquivos TXT: {e}")
                    # --- FIM DA MODIFICAÇÃO ---

                    # Calcula as somas das colunas
                    soma_cosip = df_controle['COSIP (R$)'].sum()
                    soma_d = df_controle['Energia (1,2%)'].sum()
                    soma_e = df_controle['Retenção(1,2%)'].sum()
                    soma_f = df_controle['Energia (4,8%)'].sum()
                    soma_g = df_controle['Retenção(4,8%)'].sum()

                    # Cria a linha em branco e a linha de totais
                    linha_em_branco = pd.DataFrame([ {col: '' for col in df_controle.columns} ])
                    linha_totais = pd.DataFrame([{
                        'UC': 'Totais:',
                        'Centro de Custo': '',
                        'Subseção': '',
                        'COSIP (R$)': soma_cosip,
                        'Energia (1,2%)': soma_d,
                        'Retenção(1,2%)': soma_e,
                        'Energia (4,8%)': soma_f,
                        'Retenção(4,8%)': soma_g
                    }])

                    # Concatena o DataFrame original com as novas linhas
                    df_controle = pd.concat([df_controle, linha_em_branco, linha_totais], ignore_index=True)

            else:
                self.log_message("AVISO: Nenhum dado extraído para gerar a aba 'Controle'.", "WARNING")
        
        # --- Prepara o DataFrame para a aba 'Relatorio' (apenas com as colunas originais) ---
        df_extracted_data = pd.DataFrame()
        if not df_full_data.empty:
            df_extracted_data = df_full_data.reindex(columns=final_columns_order_data)

            # Formata colunas de moeda para cálculo
            for col_name in currency_cols_names_for_excel_fmt:
                if col_name in df_extracted_data.columns:
                    df_extracted_data[col_name] = pd.to_numeric(df_extracted_data[col_name], errors='coerce').fillna(0.0)

        # --- Create TOTAL row for extracted data ('Relatorio') ---
        df_total_row = pd.DataFrame() # Initialize empty
        if not df_extracted_data.empty:
            total_row_data = {"UC": "Totais:"}
            for col in final_columns_order_data:
                if col in currency_cols_names_for_excel_fmt:
                    total_row_data[col] = df_extracted_data[col].sum()
                elif col != "UC":
                    total_row_data[col] = ""
            df_total_row = pd.DataFrame([total_row_data]).reindex(columns=final_columns_order_data)

        # --- Create Valor Cobrado summary row ---
        df_cobrado_summary_row = pd.DataFrame() # Initialize empty
        if all_valor_cobrado_results:
            total_valor_cobrado_sum = sum(res.get("liquido_total_verified", 0.0) for res in all_valor_cobrado_results if res.get("liquido_total_verified") is not None)
            
            cobrado_summary_row_data = {col: "" for col in final_columns_order_data}
            cobrado_summary_row_data["UC"] = "TOTAL conta:"
            if "LÍQUIDO (R$)" in final_columns_order_data:
                cobrado_summary_row_data["LÍQUIDO (R$)"] = total_valor_cobrado_sum
            df_cobrado_summary_row = pd.DataFrame([cobrado_summary_row_data]).reindex(columns=final_columns_order_data)
            self.log_message(f"Soma total de 'Valor Cobrado Verificado': {total_valor_cobrado_sum}", "INFO")

        # --- Assemble the final report DataFrame ('Relatorio') with blank rows ---
        final_report_parts = []
        if not df_extracted_data.empty:
            final_report_parts.append(df_extracted_data)
            if not df_total_row.empty or not df_cobrado_summary_row.empty:
                blank_row_df = pd.DataFrame([{col: "" for col in final_columns_order_data}])
                final_report_parts.append(blank_row_df)

        if not df_total_row.empty:
            final_report_parts.append(df_total_row)
        if not df_cobrado_summary_row.empty:
            final_report_parts.append(df_cobrado_summary_row)
        
        df_final_report = pd.concat(final_report_parts, ignore_index=True) if final_report_parts else pd.DataFrame(columns=final_columns_order_data)

        # --- Process df_errors ---
        df_errors = pd.DataFrame()
        if error_items:
            df_errors = pd.DataFrame(error_items).reindex(columns=final_columns_order_errors)

        # --- Save and open the Excel file ---
        try:
            self.log_message(f"Salvando relatório em: {output_file_path}", "INFO")
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                worksheet = None # Initialize worksheet variable
                if not df_final_report.empty:
                    df_final_report.to_excel(writer, index=False, sheet_name='Relatorio')
                    workbook = writer.book
                    worksheet = writer.sheets['Relatorio'] # Get worksheet here
                    worksheet.freeze_panes = 'A2' # Congela a linha de cabeçalho
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    
                    # Alinhar à direita a coluna 'Numero da Pagina'
                    try:
                        col_index = final_columns_order_data.index("Numero da Pagina") + 1
                        col_letter = get_column_letter(col_index)
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet[f"{col_letter}{row}"]
                            cell.alignment = Alignment(horizontal="right")
                    except (ValueError, IndexError):
                        pass # Ignora se a coluna não for encontrada

                    # Formata colunas de moeda e aplica destaque condicional
                    for col_name_df in currency_cols_names_for_excel_fmt:
                        if col_name_df in df_final_report.columns:
                            excel_col_idx = final_columns_order_data.index(col_name_df) + 1
                            col_letter = get_column_letter(excel_col_idx)
                            
                            for row_idx_in_final_df in range(df_final_report.shape[0]):
                                row_excel_num = row_idx_in_final_df + 2 
                                cell = worksheet[f'{col_letter}{row_excel_num}']
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = 'R$ #,##0.00'
                                    if cell.value == 0 and col_name_df in ["LÍQUIDO (R$)", "COSIP (R$)"]:
                                        cell.fill = yellow_fill

                    # Ajusta largura das colunas
                    for col_idx_df, col_name_df in enumerate(final_columns_order_data):
                        excel_col_idx = col_idx_df + 1
                        column_letter_val = get_column_letter(excel_col_idx)
                        max_len = len(str(worksheet[f'{column_letter_val}1'].value))
                        
                        for cell in worksheet[column_letter_val]:
                            if cell.value:
                                cell_str_val = str(cell.value)
                                if col_name_df in currency_cols_names_for_excel_fmt and isinstance(cell.value, (int, float)):
                                    cell_str_val = f"R$ {cell.value:,.2f}"
                                max_len = max(max_len, len(cell_str_val))
                        
                        adjusted_width = (max_len + 2) if max_len > 0 else 12
                        if col_name_df == "UC":
                             adjusted_width = max(adjusted_width, 15) 
                        worksheet.column_dimensions[column_letter_val].width = adjusted_width

                # --- GRAVAR A ABA 'CONTROLE' (SE GERADA) ---
                if not df_controle.empty:
                    df_controle.to_excel(writer, index=False, sheet_name='Controle')
                    worksheet_controle = writer.sheets['Controle']
                    worksheet_controle.freeze_panes = 'A2'

                    # Lista de colunas de moeda para a nova aba 'Controle'
                    controle_currency_cols = [
                        "COSIP (R$)",
                        "Energia (1,2%)", "Retenção(1,2%)",
                        "Energia (4,8%)", "Retenção(4,8%)"
                    ]
                    
                    # Formatar colunas para a aba 'Controle'
                    for col_idx, col_name in enumerate(df_controle.columns):
                        col_letter = get_column_letter(col_idx + 1)
                        for row_num in range(2, worksheet_controle.max_row + 1):
                            cell = worksheet_controle[f'{col_letter}{row_num}']
                            if col_idx == 0: worksheet_controle.row_dimensions[row_num].height = 15
                            # Formata colunas de moeda
                            if col_name in controle_currency_cols and isinstance(cell.value, (int, float)):
                                cell.number_format = 'R$ #,##0.00'
                            # Aplica quebra de linha na coluna UC
                            if col_name == 'UC' and cell.value and isinstance(cell.value, str) and '\n' in cell.value:
                                cell.alignment = Alignment(wrap_text=True, vertical='top')

                    # Ajustar largura das colunas para a aba 'Controle'
                    for col_idx, col_name in enumerate(df_controle.columns):
                        column_letter = get_column_letter(col_idx + 1)
                        max_len = len(str(worksheet_controle[f'{column_letter}1'].value))
                        for cell in worksheet_controle[column_letter]:
                            if cell.value:
                                cell_str = str(cell.value)
                                if col_name == 'UC':
                                    # Para a coluna UC, a largura é baseada na linha mais longa (UC mais longa)
                                    lines = cell_str.split('\n')
                                    current_max_line_len = max(len(line) for line in lines) if lines else 0
                                    max_len = max(max_len, current_max_line_len)
                                else:
                                    # Para outras colunas, usa o comprimento total da string
                                    if col_name in controle_currency_cols and isinstance(cell.value, (int, float)):
                                        cell_str = f"R$ {cell.value:,.2f}"
                                    max_len = max(max_len, len(cell_str))
                        adjusted_width = (max_len + 2) if max_len > 0 else 12
                        worksheet_controle.column_dimensions[column_letter].width = adjusted_width

                if not df_errors.empty:
                    df_errors.to_excel(writer, index=False, sheet_name='Relatorio_Erros')
                    worksheet_errors = writer.sheets['Relatorio_Erros']
                    worksheet_errors.freeze_panes = 'A2' # Congela a linha de cabeçalho
                    for col_idx_df, col_name_df in enumerate(final_columns_order_errors):
                        excel_col_idx = col_idx_df + 1
                        column_letter_val = get_column_letter(excel_col_idx)
                        max_len = len(str(worksheet_errors[f'{column_letter_val}1'].value))
                        for cell in worksheet_errors[column_letter_val]:
                             if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        adjusted_width = (max_len + 2) if max_len > 0 else 12
                        if col_name_df == "Observação":
                            adjusted_width = min(adjusted_width, 80) # Limita a largura da coluna de observação
                        worksheet_errors.column_dimensions[column_letter_val].width = adjusted_width
            
            # --- Perform the value comparison and apply highlight ---
            calculated_total_liquido = df_total_row['LÍQUIDO (R$)'].iloc[0] if not df_total_row.empty else 0.0
            account_total_liquido = df_cobrado_summary_row['LÍQUIDO (R$)'].iloc[0] if not df_cobrado_summary_row.empty else 0.0

            if abs(calculated_total_liquido - account_total_liquido) > 1e-9:
                self.account_values_mismatched = True
                self.log_message("Valores da conta não conferem! (Total Extraído vs Total da Fatura)", "WARNING")

                if worksheet is not None:
                    totais_row_index_in_sheet = -1
                    for r_idx in range(2, worksheet.max_row + 1):
                        if worksheet[f'A{r_idx}'].value == "Totais":
                            totais_row_index_in_sheet = r_idx
                            break

                    if totais_row_index_in_sheet != -1:
                        col_name_to_highlight = "LÍQUIDO (R$)"
                        if col_name_to_highlight in final_columns_order_data:
                            excel_col_idx_highlight = final_columns_order_data.index(col_name_to_highlight) + 1
                            col_letter_highlight = get_column_letter(excel_col_idx_highlight)
                            
                            worksheet[f'{col_letter_highlight}{totais_row_index_in_sheet}'].fill = yellow_fill
                            self.log_message(f"Célula {col_letter_highlight}{totais_row_index_in_sheet} (Totais, LÍQUIDO) destacada em amarelo.", "INFO")
                    else:
                        self.log_message("AVISO: Não foi possível localizar a linha 'Totais' para destacar o valor.", "WARNING")

            # --- Determine Final Status and Messages ---
            final_status_message = ""
            final_messagebox_title = ""
            final_messagebox_type = messagebox.showinfo
            summary_message = ""

            if self.account_values_mismatched:
                final_status_message = "Concluído: Valores da conta não conferem!"
                final_messagebox_title = "Alerta Crítico: Discrepância nos Valores!"
                summary_message = (f"ATENÇÃO: Os valores totais calculados e os valores informados na conta não conferem.\n"
                                   f"Verifique a linha 'Totais' na aba 'Relatorio' (destacada em amarelo).\n"
                                   f"Total Calculado: R$ {calculated_total_liquido:,.2f}\n"
                                   f"Total da Fatura: R$ {account_total_liquido:,.2f}\n\n")
                final_messagebox_type = messagebox.showwarning
                if self.current_severity < 1:
                    self.current_severity = 1
            elif self.current_severity == 2:
                final_status_message = "Concluído com ERROS."
                final_messagebox_title = "Processamento Concluído com Alertas"
                summary_message = f"Processamento concluído com ERROS!\n"
                if not df_extracted_data.empty:
                    summary_message += f"{len(df_extracted_data)} registros de fatura extraídos com sucesso na aba 'Relatorio'.\n"
                summary_message += f"{len(df_errors)} problemas/erros encontrados na aba 'Relatorio_Erros'."
                final_messagebox_type = messagebox.showerror
            elif self.has_specific_warnings:
                final_status_message = "Concluído com Avisos!"
                final_messagebox_title = "Processamento Concluído com Avisos"
                summary_message = f"Processamento concluído com Avisos!\n"
                if not df_extracted_data.empty:
                    summary_message += f"{len(df_extracted_data)} registros de fatura extraídos na aba 'Relatorio'.\n"
                final_messagebox_type = messagebox.showwarning
            elif df_final_report.empty:
                final_status_message = "Concluído (Sem dados extraídos)."
                final_messagebox_title = "Processamento Concluído"
                summary_message = ("Processamento concluído. Nenhum dado de fatura válido foi extraído.\n"
                                   "Verifique se os PDFs contêm faturas com UCs identificáveis.")
                final_messagebox_type = messagebox.showinfo
            else:
                final_status_message = "Concluído com sucesso!"
                final_messagebox_title = "Processamento Concluído"
                summary_message = f"Processamento concluído com sucesso!\n{len(df_extracted_data)} registros de fatura extraídos na aba 'Relatorio'."
                final_messagebox_type = messagebox.showinfo
            
            final_progress_bar_style = "Success.Horizontal.TProgressbar"
            if self.current_severity == 1:
                final_progress_bar_style = "Warning.Horizontal.TProgressbar"
            elif self.current_severity == 2:
                final_progress_bar_style = "Error.Horizontal.TProgressbar"
            self.set_progress_bar_style(final_progress_bar_style)

            self.status_label.config(text=final_status_message)

            if output_file_path:
                final_summary_msg_for_box = summary_message + f"\nRelatório salvo em:\n{output_file_path}"
                final_messagebox_type(final_messagebox_title, final_summary_msg_for_box)

        except Exception as e:
            self.log_message(f"Erro CRÍTICO ao salvar o relatório Excel: {e}", "CRITICAL_ERROR")
            messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar o relatório: {e}")
            self.status_label.config(text="Erro ao salvar relatório.")
        finally:
            self.process_button.config(state=tk.NORMAL)
            
            if os.path.exists(output_file_path):
                try:
                    if sys.platform == "win32":
                        os.startfile(output_file_path)
                    elif sys.platform == "darwin":
                        subprocess.call(("open", output_file_path))
                    else:
                        subprocess.call(("xdg-open", output_file_path))
                except Exception as open_e:
                    self.log_message(f"Não foi possível abrir o relatório automaticamente: {open_e}", "WARNING")
            else:
                self.log_message(f"AVISO: Arquivo de relatório não encontrado para abrir: {output_file_path}", "WARNING")


    # --- FUNÇÃO ADICIONADA: Extrair e Verificar Valor Cobrado ---
    def clean_currency(self, value_str):
        """Limpia uma string de valor monetário (ex: 1.234,56) para float (ex: 1234.56)."""
        if not isinstance(value_str, str) or not value_str.strip():
            return None
        # Remove todos os pontos e substitui a vírgula por ponto para conversão para float
        cleaned_str = value_str.strip().replace('.', '').replace(',', '.')
        try:
            return float(cleaned_str)
        except ValueError:
            return None # Retorna None se não puder converter

    def extract_and_verify_valor_cobrado(self, pdf_path):
        """
        Extrai o 'Valor Cobrado' e verifica sua duplicação na primeira página do PDF.
        Retorna o valor cobrado, sua string original, o líquido total (se duplicado)
        e uma lista de mensagens de status.
        """
        valor_cobrado = None
        valor_cobrado_str_original = None
        liquido_total = None # Representa o valor cobrado verificado
        status_messages = []

        try:
            # 1 Abrir o PDF e acessar a primeira página
            doc = fitz.open(pdf_path)
            if doc.page_count == 0:
                doc.close()
                return None, None, None, ["Erro: O PDF não contém páginas."]
            
            page = doc[0] # Foca apenas na primeira página
            text = page.get_text("text") # Extrai todo o texto da primeira página
            doc.close() # Fecha o documento

            # 2 Procurar o rótulo "Valor Cobrado (R$)"
            match_label_cobrado = re.search(r"Valor Cobrado \(R\$\)", text, re.IGNORECASE)

            if match_label_cobrado:
                # 3 Definir uma área de busca restrita após o rótulo
                search_start = match_label_cobrado.end()
                # Limita a busca aos próximos 150 caracteres após o rótulo (ajustável)
                search_end = min(search_start + 150, len(text))
                search_text_area = text[search_start:search_end]

                # 4 Extrair o(s) valor(es) numérico(s) da área de busca
                potential_values = re.findall(r"([\d.,]+)", search_text_area)

                if potential_values:
                    # 5 Limpar e converter as strings de valores para números reais
                    # 6 Pegar o primeiro valor numérico encontrado
                    for val_str in potential_values:
                        cleaned_val = self.clean_currency(val_str)
                        if cleaned_val is not None:
                            valor_cobrado_str_original = val_str # Guarda a string original
                            valor_cobrado = cleaned_val
                            status_messages.append(f"Encontrado 'Valor Cobrado': '{valor_cobrado_str_original}' -> {valor_cobrado}")
                            break # Para no primeiro valor válido encontrado

                    if valor_cobrado is None:
                        status_messages.append("Aviso: Nenhum valor numérico válido encontrado após 'Valor Cobrado (R$)'.")
                else:
                    status_messages.append("Aviso: Nenhum valor numérico encontrado na área de busca após 'Valor Cobrado (R$)'.")
            else:
                status_messages.append("Aviso: Rótulo 'Valor Cobrado (R$)' não encontrado na primeira página.")

            # 7 Verificar a duplicação na primeira página
            if valor_cobrado is not None and valor_cobrado_str_original is not None:
                # Busca TODAS as ocorrências da STRING original do valor cobrado no texto COMPLETO da primeira página
                all_occurrences = list(re.finditer(re.escape(valor_cobrado_str_original), text))

                # Conta as ocorrências
                if len(all_occurrences) >= 2: # Se o valor aparece 2 vezes ou mais
                    # 9 Salvar o resultado: Se duplicado
                    liquido_total = valor_cobrado # Confirma o valor que será somado
                    status_messages.append(f"Sucesso: Valor '{valor_cobrado_str_original}' encontrado {len(all_occurrences)} vezes na página.")
                else:
                    status_messages.append("Falha: O valor do 'Valor Cobrado' não foi encontrado duplicado na página.")
            elif valor_cobrado is None:
                 status_messages.append("Falha: Não é possível verificar duplicação pois 'Valor Cobrado' não foi extraído.")

            return valor_cobrado, valor_cobrado_str_original, liquido_total, status_messages

        except Exception as e:
            # Captura qualquer erro inesperado durante o processamento
            status_messages.append(f"Erro inesperado durante a extração/verificação: {e}")
            return None, None, None, status_messages

    def show_info(self):
        """
        Abre um pop-up com informações sobre o programa.
        """
        info_popup = Toplevel(self.root)
        info_popup.title("Informação")
        info_popup.transient(self.root)
        info_popup.grab_set()
        info_popup.resizable(False, False)
        info_popup.configure(bg="#f0f0f0")

        if os.path.exists(self.icon_path):
            try:
                info_popup.iconbitmap(self.icon_path)
            except tk.TclError as e:
                print(f"Erro ao carregar ícone para o popup: {e}")

        content_frame = Frame(info_popup, padx=15, pady=15, bg=info_popup.cget("bg"))
        content_frame.pack(expand=True, fill=tk.BOTH)

        version_label = Label(content_frame, text=f"{self.root.title()} - by Elias", font=("Segoe UI", 10), bg=content_frame.cget("bg"), fg="#002b00")
        version_label.pack(pady=(0,5))
        pix_label = Label(content_frame, text="Chamado via mensagem Pix: eliasgkersten@gmail.com", font=("Segoe UI", 10), bg=content_frame.cget("bg"), fg="#002b00")
        pix_label.pack(pady=5)

        close_button = ttk.Button(content_frame, text="OK", command=info_popup.destroy)
        close_button.pack(pady=10)

        info_popup.update_idletasks()
        popup_width = info_popup.winfo_width()
        popup_height = info_popup.winfo_height()
        self.center_window_for_popup(info_popup, popup_width, popup_height)

    def center_window_for_popup(self, window_to_center, width, height):
        """Centraliza uma janela (como um popup) na tela."""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        window_to_center.geometry(f'{width}x{height}+{int(x)}+{int(y)}')


# Função para criar botões arredondados (mantida como função externa conforme original)
def create_rounded_button(parent, text, command, width=20, height=20, bg_color=None):
    canvas_bg = bg_color if bg_color else parent.cget("bg")
    canvas = Canvas(parent, width=width, height=height, bd=0, highlightthickness=0, relief='ridge', bg=canvas_bg)
    canvas.create_oval(1, 1, width-2, height-2, outline="#0000FF", fill="#0000FF")
    canvas.create_text(width/2, height/2, text=text, fill="#FFFFFF", font=("Segoe UI Bold", int(height/2)))
    canvas.bind("<Button-1>", lambda event: command())
    return canvas

if __name__ == "__main__":
    root = tk.Tk()
    app = AppCelescReporter(root)
    root.mainloop()